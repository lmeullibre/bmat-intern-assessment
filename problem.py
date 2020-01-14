import openpyxl
import csv
import pandas as pd
import spotipy
from spotipy import oauth2


scope = 'user-library-read'

sp_oauth = oauth2.SpotifyOAuth(client_id='6a97984e552942e6808c13e6562543bf',
                                  client_secret='c31496d96a3a47278cf6be88bb54d013',
                                  redirect_uri='https://www.google.es',
                                  scope=scope)


def refresh():
    global token_info, sp
    token_info = sp_oauth.refresh_access_token(token_info['refresh_token'])
    token = token_info['access_token']
    print("refreshing spotify, new token:" + token)
    sp = spotipy.Spotify(auth=token)
    return sp


def getLongSpotify():
    global token_info, sp
    token_info = sp_oauth.get_cached_token()
    if not token_info:
        auth_url = sp_oauth.get_authorize_url()
        print(auth_url)
        response = input('Paste the above link into your browser, then paste the redirect url here: (This is for all the Spotify stuff) ')

        code = sp_oauth.parse_response_code(response)
        token_info = sp_oauth.get_access_token(code)

        token = token_info['access_token']

    sp = spotipy.Spotify(auth=token)
    return sp



def grabSong(artist, title, spotify):
    resultats = spotify.search(q='artist:' + artist + ' track:' + title, type='track')
    trackId = resultats['tracks']['items'][0]['id']
    return trackId


def getSongId(artist, title, songMap, spotify):
    index = artist + title
    trackId = 'null'

    try:
        trackId = songMap[index]
    except KeyError:
        try:
            trackId = grabSong(artist, title, spotify)
            songMap[index] = trackId
        except IndexError:
            trackId = 'null'

    return trackId, songMap


def stringhourToDate(hora):  # turns a string that reprents an hour into a date object
    from datetime import datetime
    h = datetime.strptime(hora, '%H:%M:%S')
    return h


def changeFormat(str):  # turns date dd/mm/yyyy into dd-mm-yyyy
    s = list(str)
    s[2] = '-'
    s[5] = '-'
    dataString = ''.join(s)
    return dataString


def formatejar(data,hora):  # turns a string that represents a date plus a string that represents an hour into an actual datetype object
    from datetime import datetime
    dataString = changeFormat(data)
    date = dataString + ' ' + hora
    datetime_object = datetime.strptime(date, '%d-%m-%Y %H:%M:%S')
    return datetime_object


def computeFinalDate(data, duration):  #
    from datetime import timedelta
    d = stringhourToDate(duration)
    segons = d.second + d.minute * 60 + d.hour * 3600
    return data + timedelta(seconds=segons)


# we initialize all the necessary buffers
book = openpyxl.load_workbook('bmat-broadcast-sw-test-epg.xlsx', data_only=True)
sheet = book.active
escritura = open('output.csv', mode='w', newline='', encoding="utf-8")
escritura = csv.writer(escritura, delimiter=',', quotechar='"')
panda = pd.read_csv("bmat-broadcast-sw-test-identifications.csv", dtype='str')
panda['start_time'] = pd.to_datetime(panda['start_time'])
panda['end_time'] = pd.to_datetime(panda['end_time'])

spotify = getLongSpotify()

songMap = {}

firsTime = True
i = 0
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row-1):

    programStartTime = formatejar(row[2].value, row[3].value)
    programEndTime = computeFinalDate(programStartTime, row[4].value)
    programName = row[1].value

    stop = False
    while not stop and i < panda.size:
        innerAux = panda.iloc[i]

        songStartTime = innerAux['start_time']
        songEndTime = innerAux['end_time']
        artistName = innerAux['artist']
        songName = innerAux['title']
        spotifyId, songMap = getSongId(artistName, songName, songMap, spotify)  # VERY TIME CONSUMING
        trackId = innerAux['id']

        if i%20000 == 0: spotify = refresh() #the connection with spotify closes after 1 hour. Couldn't manage to refresh the authorization via exception handling

        if (songStartTime < programStartTime):
            songStartTime = programStartTime
        if (songEndTime > programEndTime):
            stop = True
        if not stop:
            escritura.writerow([programName, songStartTime, songEndTime, artistName, songName, spotifyId])
            print(trackId)
            i = i + 1

# handling last row
lastRow = panda.iloc[-1]
lR = sheet[sheet.max_row]
id, songMap = getSongId(lastRow['artist'], lastRow['title'], songMap, spotify)
escritura.writerow([lR[1].value, lastRow['start_time'], lastRow['end_time'], lastRow['artist'], lastRow['title'], id ])